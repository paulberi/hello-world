# This Python script reads information about layers, layer groups and roles from a given customer 
# appConfig and writes them into an Excel file in an organized way. It can also deletes a given
# customer's layers and roles from the Excel file.
# --- Reza Firouzfar 2021-07-23 ---

import os
import psycopg2
import pandas as pd
import xlwings as xw
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook  
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# ----------- Replace ÅÄÖåäö ----------
def letter_decoder(text):
  text = text.replace("Ã…","Å").replace("Ã„","Ä").replace("Ã–","Ö").replace("Ã‰","É").replace("Ãœ","Ü").replace("Ã€","À").replace("Ã¥","å").replace("Ã¤","ä").replace("Ã¶","ö").replace("Ã©","é").replace("Ã¼","ü").replace("Ã","à")
  return text

# 
actionType = ""
while (len(actionType) == 0):
  actionType = input('Do you want to add customer to the Excel file or delete customer? [add/del]: ')
  actionType = actionType.strip()
  actionType = actionType.lower()
  if not actionType == "add" and not actionType == "del":
    actionType = ""
if actionType == 'add':  
  # ----------- ask for source file name ----------
  sourceFileName = ""
  while (len(sourceFileName) == 0):
    sourceFileName = input('Please input the source file name (json): ')
    sourceFileName =  sourceFileName.strip()
    sourceFileName = letter_decoder(sourceFileName)

# ----------- ask for destination file name ----------
destinationFileName = ""
while (len(destinationFileName) == 0):
  destinationFileName = input('Please input the destination file name (xlsx): ')
  destinationFileName = destinationFileName.strip()
  destinationFileName = letter_decoder(destinationFileName)

# ---------- create the output file, asks question if file exists ----------
if os.path.exists(destinationFileName):
  answer = input('Are you sure to update "' + destinationFileName + '? [y/n]: ')
  if answer == 'y':
    wb = load_workbook(destinationFileName)
  else:
    print("Interupted!")
    exit()
else:
  wb = load_workbook(destinationFileName)

customerName = ""
while (len(customerName) == 0):
  customerName = input('Please input the exact customer name: ')
  customerName = letter_decoder(customerName.strip())

geovisType = ""
while (len(geovisType) == 0):
  geovisType = input('Please input the kind of GeoVis, (publik, vy or pro): ')
  geovisType = geovisType.strip()
  geovisType = geovisType.lower()
  if not geovisType == "publik" and not geovisType == "vy" and not geovisType == "pro":
    geovisType = ""

# ---------- Replace group Name with group title ----------
def changeGroupNameToGroupTitle(groupName, groupNameList, groupTitleList):
  if groupName == "" or groupName is None:
    return ""
  groupNameLength = len(groupName)
  m = -1
  for i in groupNameList:
    m = m +1
    if i.find(groupName) != -1:
      theName = i
      theNameLength = len(theName)
      x = theNameLength - groupNameLength
      y = theNameLength
      if theName[x:y] == groupName:
        return groupTitleList[m]
#    else:
#      print("Not find")

# ----------- Delete empty rows in Excel sheet -------
def deleteEmptyRows(destinationFileName, workBook, sheet):
  rows_to_delete = [None, '', ' ']
  # Following three lines is used if you would like to delete empty rows in all sheets. 
#  for i in workBook.sheetnames:
#    print(f'Now in sheet: {i}')
#    sheet = workBook[i]
  # loop each row in column A, delete row if column A is empty.
  column_a = range(1, sheet.max_row)
  for i in reversed(column_a):
    if sheet.cell(i, 1).value in rows_to_delete:
#      print(f'Deleting Row: {sheet.cell(i,1).row}')
      sheet.delete_rows(sheet.cell(i,1).row)

  if sheet.cell(sheet.max_row, 1).value in rows_to_delete:
#    print(f'Deleting Row: {sheet.cell(sheet.max_row,1).row}')
    sheet.delete_rows(sheet.cell(sheet.max_row,1).row)
  workBook.save(destinationFileName)

# ---------- WashString ----------
def washString(text):
  text = text.strip()
  startPos = -1
  separator = ":"
  for i in range(0, len(text)-1):
    if text[i] == separator:
      startPos = i+1
      break

  if startPos > -1:  
    text = text[startPos:]  
    startPos = text.index('"') 
    text = text[startPos + 1:]
    startPos = text.index('"')
    text = text[0:startPos]
    return letter_decoder(text)

# ----------- Delete Excel rows belonging a customer name from a given sheet and roleSheet ----------
def deleteCustomerFromExcelSheet(destinationFileName, workBook, sheet, roleSheet, name):
  if actionType == "del":
    sheetList = [sheet, roleSheet]
    # loop each row in column B, delete rows with matching posts in column B.
    for x in sheetList:
      column_b = range(1, x.max_row)
      for i in reversed(column_b):
        if x.cell(i, 2).value == name:
          print(f'Deleting Row: {x.cell(i,2).row}')
          x.delete_rows(x.cell(i,2).row)
      if x.cell(x.max_row, 2).value == name:
          print(f'Deleting Last Row: {x.cell(x.max_row,2).row}')
          x.delete_rows(x.cell(x.max_row,2).row)

    workBook.save(destinationFileName)
    print(destinationFileName + " was updated!")
    exit()

# ---------- Update Excel sheet "Kundinformation", the first sheet ----------
def updateCustomerInfoSheet(destinationFileName, workBook, sheet):
  sheet.cell(row = 1, column = 1).font = Font(bold=True)
  sheet.cell(row = 1, column = 1).value = "kund"
  sheet.cell(row = 1, column = 2).font = Font(bold=True)
  sheet.cell(row = 1, column = 2).value = "Kontaktpersoner"
  sheet.cell(row = 1, column = 3).font = Font(bold=True)
  sheet.cell(row = 1, column = 3).value = "E-post"
  sheet.cell(row = 1, column = 4).font = Font(bold=True)
  sheet.cell(row = 1, column = 4).value = "Tel/Mobil"
  sheet.cell(row = 1, column = 5).font = Font(bold=True)
  sheet.cell(row = 1, column = 5).value = "Produkt"
  sheet.cell(row = 1, column = 6).font = Font(bold=True)
  sheet.cell(row = 1, column = 6).value = "Avtal"
  sheet.cell(row = 1, column = 7).font = Font(bold=True)
  sheet.cell(row = 1, column = 7).value = "URL till applikation"
  sheet.cell(row = 1, column = 8).font = Font(bold=True)
  sheet.cell(row = 1, column = 8).value = "Koordinatsystem"
  sheet.cell(row = 1, column = 9).font = Font(bold=True)
  sheet.cell(row = 1, column = 9).value = "Geokodningstjänst"
  sheet.cell(row = 1, column = 10).font = Font(bold=True)
  sheet.cell(row = 1, column = 10).value = "Konto Metria Maps (user:pwd)"    
  sheet.cell(row = 1, column = 11).font = Font(bold=True)
  sheet.cell(row = 1, column = 11).value = "Konto Geoassistans (user:pwd)"  
  sheet.cell(row = 1, column = 12).font = Font(bold=True)
  sheet.cell(row = 1, column = 12).value = "Konto andra datakällor (user:pwd)"  
  sheet.cell(row = 1, column = 13).font = Font(bold=True)
  sheet.cell(row = 1, column = 13).value = "Xplore DB schema"  
  sheet.cell(row = 1, column = 14).font = Font(bold=True)
  sheet.cell(row = 1, column = 14).value = "Geografisk begränsning"
  sheet.cell(row = 1, column = 15).font = Font(bold=True)
  sheet.cell(row = 1, column = 15).value = "Behörighetsnivåer"
  sheet.cell(row = 1, column = 16).font = Font(bold=True)
  sheet.cell(row = 1, column = 16).value = "Kommentarer"
  workBook.save(destinationFileName)


sheets = wb.sheetnames
if geovisType == "publik":
  destinationSheet = wb[sheets[1]]
elif geovisType == "vy":
  destinationSheet = wb[sheets[2]]
else:
  destinationSheet = wb[sheets[3]]
customerInfoSheet = wb[sheets[0]]
roleDescriptionSheet = wb[sheets[4]]

updateCustomerInfoSheet(destinationFileName, wb, customerInfoSheet)
deleteEmptyRows(destinationFileName, wb, destinationSheet)
deleteEmptyRows(destinationFileName, wb, roleDescriptionSheet)
deleteCustomerFromExcelSheet(destinationFileName, wb, destinationSheet, roleDescriptionSheet, customerName) 

# ---------- Read source file and update Excel file ----------
with open(sourceFileName) as f1:

# ------- For groups --------
  groupLevel_1_exists = 0
  groupLevel_2_exists = 0
  groupLevel_3_exists = 0
  groupNameList = []
  groupTitleList = []
  nameExists = 0
  titleExists = 0
  groupNameLevel_1= ""
  groupTitleLevel_1 = ""
  groupNameLevel_2= ""
  groupTitleLevel_2 = ""
  groupNameLevel_3= ""
  groupTitleLevel_3 = ""
  tempGroupName = ""
  tempGroupTitle = ""
  layerDetected = 0

# ------- For Layers --------
  layerDetected = 0
  defaultUrlDetected = 0
  defaultUrl = ""
  rowNr = 0
  postNr = 1
  rowUpdated = 0
  nr_of_wms_layers = 0
  nr_of_wfs_layers = 0
  temp_group = ""
  temp_title = ""
  temp_serverType = ""
  temp_allowedRoles = ""
  temp_allowedRolesTemplate = ""
  temp_url = ""
  temp_attribute = ""
  splited_wms_line = ""
  splited_wfs_line = ""
  wms_layer = ""
  wfs_layer = ""

# ------- For Roles --------
  roleRowUpdated = 0
  roleRowNr = 0
  rolePostNr = 1
  role_group = ""
  roleList = []
  roleDetected = 0


  line = f1.readline()
  print("Reading layer groups ....")
  while line:
    # ---------------------- Read groups ------------------------ 
    if defaultUrlDetected == 0:
      if '"defaultUrls":' in line:
        defaultUrlDetected = 1
    if defaultUrlDetected == 1:
      if '"url":' in line:
        defaultUrl = washString(line)
        defaultUrl = defaultUrl + " (Xplore)"
        defaultUrlDetected = 0

    if layerDetected == 0:
      if '"groups":' in line:
        pos = line.find(":")
        if pos == 10:
          groupLevel_1_exists = 1
          groupNameLevel_1 = ""
          groupTitleLevel_1 = ""
        elif pos == 14:
          groupLevel_2_exists = 1
          groupNameLevel_2 = ""
          groupTitleLevel_2 = ""
        elif pos == 18:
          groupLevel_3_exists = 1
          groupNameLevel_3 = ""
          groupTitleLevel_3 = ""

      if '  ]' in line:
        pos = line.find("]")
        if pos == 2:
          groupLevel_1_exists = 0
        elif pos == 6:
          groupLevel_2_exists = 0
        elif pos == 10:
          groupLevel_3_exists = 0

      if groupLevel_1_exists == 1:
        if not '    },' == line.rstrip():
          if '"name":' in line:
            tempGroupName = washString(line)
            nameExists = 1
          if '"title":' in line:
            tempGroupTitle = washString(line)
            titleExists = 1

          if groupLevel_3_exists == 1:
            groupNameLevel_3 = tempGroupName
            groupTitleLevel_3 = tempGroupTitle
          elif groupLevel_2_exists == 1:
            groupNameLevel_2 = tempGroupName
            groupTitleLevel_2 = tempGroupTitle
          elif groupLevel_1_exists == 1:
            groupNameLevel_1 = tempGroupName
            groupTitleLevel_1 = tempGroupTitle

          if nameExists == 1 and titleExists == 1:

            if groupLevel_3_exists == 1:
              groupNameList.append(groupNameLevel_1 + "/" + groupNameLevel_2 + "/" + groupNameLevel_3)
              groupTitleList.append(groupTitleLevel_1 + "/" + groupTitleLevel_2 + "/" + groupTitleLevel_3)

            elif groupLevel_2_exists == 1:
              groupNameList.append(groupNameLevel_1 + "/" + groupNameLevel_2)
              groupTitleList.append(groupTitleLevel_1 + "/" + groupTitleLevel_2)
              groupNameLevel_3 = ""
              groupTitleLevel_3 = ""
            elif groupLevel_1_exists == 1:
              groupNameList.append(groupNameLevel_1)
              groupTitleList.append(groupTitleLevel_1)
              groupNameLevel_2 = ""
              groupTitleLevel_2 = ""
            nameExists = 0
            titleExists = 0

#            print("groupNameList: ")
#            print(groupNameList)
#            print("groupTitleList: ")
#            print(groupTitleList)

        else:
          groupNameLevel_1= ""
          groupTitleLevel_1 = ""

# ---------- Read information about layers and update Excel ----------
    if '"layers": [' in line:
      print("Reading layers....")
      destinationSheet.cell(row = 1, column = 1).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 1).value = "Nr"
      destinationSheet.cell(row = 1, column = 2).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 2).value = "Kund"
      destinationSheet.cell(row = 1, column = 3).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 3).value = "Grupp"
      destinationSheet.cell(row = 1, column = 4).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 4).value = "Lager"
      destinationSheet.cell(row = 1, column = 5).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 5).value = "WMS-lager"
      destinationSheet.cell(row = 1, column = 6).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 6).value = "WFS-lager"
      destinationSheet.cell(row = 1, column = 7).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 7).value = "Attribut"
      destinationSheet.cell(row = 1, column = 8).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 8).value = "Roller"
      destinationSheet.cell(row = 1, column = 9).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 9).value = "Server type"
      destinationSheet.cell(row = 1, column = 10).font = Font(bold=True)
      destinationSheet.cell(row = 1, column = 10).value = "URL/Datakälla"

      layerDetected = 1

      if destinationSheet.max_row > 1:
        rowNr = destinationSheet.max_row + 1
      else:
        rowNr = 2

    if layerDetected == 1:
      if '"group":' in line:
        temp_group = washString(line)
        temp_group = changeGroupNameToGroupTitle(temp_group, groupNameList, groupTitleList)
        rowUpdated = 1
      if '"title":' in line:
        temp_title = washString(line)
        rowUpdated = 1
      if '"LAYERS":' in line:
        line = washString(line)

        # --------- checks if there are more than one lager in the LAYERS-list ---------
        if ',' in line:
          splited_wms_line = line.split(",")
          nr_of_wms_layers = len(splited_wms_line)
        elif line == "":
          line = "(utveckling pågår!)"
          nr_of_wms_layers = 1
          wms_layer = line
        else:
          nr_of_wms_layers = 1
          wms_layer = line
        rowUpdated = 1
      if '"typeName":' in line:
        line = washString(line)
        # --------- checks if there are more than one lager in the typeName-list ---------
        if ',' in line:
          splited_wfs_line = line.split(",")
          nr_of_wfs_layers = len(splited_wfs_line)
        elif line == "":
          nr_of_wms_layers = 1
          wms_layer = "(uyveckling pågår)"
        else:
          nr_of_wfs_layers = 1
          wfs_layer = line
        rowUpdated = 1
      if '"serverType":' in line:
        temp_serverType = washString(line)
        rowUpdated = 1
      if '"queryable":' in line:
        if "true" in line:   
          temp_attribute = "Ja"
        else:
          temp_attribute = ""
        rowUpdated = 1
      if '"allowedRoles":' in line:
        temp_allowedRoles = washString(line)
        rowUpdated = 1
      if '"allowedRolesTemplate":' in line:
        temp_allowedRolesTemplate = washString(line)
        rowUpdated = 1
      if '"url":' in line:
        temp_url = washString(line)
        if temp_title.lower() == "kommungräns":
          temp_url = temp_url + " (Xplore)"
        rowUpdated = 1
      
      if ("    }," == line.rstrip() or "  ]," == line.rstrip()) and rowUpdated == 1:
        if nr_of_wms_layers > 0:
          if nr_of_wms_layers == 1:
            wms_layer = letter_decoder(wms_layer)
            destinationSheet.cell(row = rowNr, column = 1).alignment = Alignment(horizontal='left')
            destinationSheet.cell(row = rowNr, column = 1).value = postNr
            destinationSheet.cell(row = rowNr, column = 2).value = customerName
            destinationSheet.cell(row = rowNr, column = 3).value = temp_group
            destinationSheet.cell(row = rowNr, column = 4).value = temp_title
            destinationSheet.cell(row = rowNr, column = 5).value = wms_layer
            destinationSheet.cell(row = rowNr, column = 7).value = temp_attribute
            if not temp_allowedRoles == "" and not temp_allowedRolesTemplate == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles + ", " + temp_allowedRolesTemplate
            elif not temp_allowedRoles == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles
            elif not temp_allowedRolesTemplate == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRolesTemplate
            destinationSheet.cell(row = rowNr, column = 9).value = temp_serverType
            if temp_url == "":
              temp_url = "maps.metria.se/geoserver/"
            destinationSheet.cell(row = rowNr, column = 10).value = temp_url
            postNr = postNr + 1
            rowNr = rowNr + 1

            rowUpdated = 0
            nr_of_wms_layers = 0
            temp_group = ""
            temp_title = ""
            temp_serverType = ""
            temp_allowedRoles = ""
            temp_allowedRolesTemplate = ""
            temp_url = ""
            temp_attribute = ""
            splited_wms_line = ""
            wms_layer = ""
          else:
            for x in range(nr_of_wms_layers):
              if len(splited_wms_line[x]) > 0:
                layer = splited_wms_line[x].replace('"', '')
                layer = letter_decoder(layer)
                destinationSheet.cell(row = rowNr, column = 1).alignment = Alignment(horizontal='left')
                destinationSheet.cell(row = rowNr, column = 1).value = postNr
                destinationSheet.cell(row = rowNr, column = 2).value = customerName
                destinationSheet.cell(row = rowNr, column = 3).value = temp_group
                destinationSheet.cell(row = rowNr, column = 4).value = temp_title
                destinationSheet.cell(row = rowNr, column = 5).value = layer
                destinationSheet.cell(row = rowNr, column = 7).value = temp_attribute
                if not temp_allowedRoles == "" and not temp_allowedRolesTemplate == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles + ", " + temp_allowedRolesTemplate
                elif not temp_allowedRoles == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles
                elif not temp_allowedRolesTemplate == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRolesTemplate

                destinationSheet.cell(row = rowNr, column = 9).value = temp_serverType
                if temp_url == "":
                  temp_url = "maps.metria.se/geoserver/"
                destinationSheet.cell(row = rowNr, column = 10).value = temp_url
                postNr = postNr + 1
                rowNr = rowNr + 1

            rowUpdated = 0
            nr_of_wms_layers = 0
            temp_group = ""
            temp_title = ""
            temp_serverType = ""
            temp_allowedRoles = ""
            temp_allowedRolesTemplate = ""
            temp_url = ""
            temp_attribute = ""
            splited_wms_line = ""
            wms_layer = ""

        elif nr_of_wfs_layers > 0:
          if nr_of_wfs_layers == 1:
            wfs_layer = letter_decoder(wfs_layer)
            destinationSheet.cell(row = rowNr, column = 1).alignment = Alignment(horizontal='left')
            destinationSheet.cell(row = rowNr, column = 1).value = postNr
            destinationSheet.cell(row = rowNr, column = 2).value = customerName
            destinationSheet.cell(row = rowNr, column = 3).value = temp_group
            destinationSheet.cell(row = rowNr, column = 4).value = temp_title
            destinationSheet.cell(row = rowNr, column = 6).value = wfs_layer
            destinationSheet.cell(row = rowNr, column = 7).value = temp_attribute
            if not temp_allowedRoles == "" and not temp_allowedRolesTemplate == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles + ", " + temp_allowedRolesTemplate
            elif not temp_allowedRoles == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles
            elif not temp_allowedRolesTemplate == "":
              destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRolesTemplate

            destinationSheet.cell(row = rowNr, column = 9).value = temp_serverType
            if temp_url == "":
              if defaultUrl == "":
                temp_url = "maps.metria.se/geoserver/"
              else:
                temp_url = defaultUrl 
            destinationSheet.cell(row = rowNr, column = 10).value = temp_url
            postNr = postNr + 1
            rowNr = rowNr + 1

            rowUpdated = 0
            nr_of_wfs_layers = 0
            temp_group = ""
            temp_title = ""
            temp_serverType = ""
            temp_allowedRoles = ""
            temp_allowedRolesTemplate = ""
            temp_url = ""
            temp_attribute = ""
            splited_wfs_line = ""
            wfs_layer = ""
          else:
            for x in range(nr_of_wfs_layers):
              if len(splited_wfs_line[x]) > 0:
                layer = splited_wfs_line[x].replace('"', '')
                layer = letter_decoder(layer)
                destinationSheet.cell(row = rowNr, column = 1).alignment = Alignment(horizontal='left')
                destinationSheet.cell(row = rowNr, column = 1).value = postNr
                destinationSheet.cell(row = rowNr, column = 2).value = customerName
                destinationSheet.cell(row = rowNr, column = 3).value = temp_group
                destinationSheet.cell(row = rowNr, column = 4).value = temp_title
                destinationSheet.cell(row = rowNr, column = 6).value = layer
                destinationSheet.cell(row = rowNr, column = 7).value = temp_attribute
                if not temp_allowedRoles == "" and not temp_allowedRolesTemplate == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles + ", " + temp_allowedRolesTemplate
                elif not temp_allowedRoles == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles
                elif not temp_allowedRolesTemplate == "":
                  destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRolesTemplate

                destinationSheet.cell(row = rowNr, column = 9).value = temp_serverType
                if temp_url == "":
                  if defaultUrl == "":
                    temp_url = "maps.metria.se/geoserver/"
                  else:
                    temp_url = defaultUrl 
                destinationSheet.cell(row = rowNr, column = 10).value = temp_url
                postNr = postNr + 1
                rowNr = rowNr + 1

            rowUpdated = 0
            nr_of_wfs_layers = 0
            temp_group = ""
            temp_title = ""
            temp_serverType = ""
            temp_allowedRoles = ""
            temp_allowedRolesTemplate = ""
            temp_url = ""
            temp_attribute = ""
            splited_wfs_line = ""
            wfs_layer = ""
        else:
          destinationSheet.cell(row = rowNr, column = 1).alignment = Alignment(horizontal='left')
          destinationSheet.cell(row = rowNr, column = 1).value = postNr
          destinationSheet.cell(row = rowNr, column = 2).value = customerName
          destinationSheet.cell(row = rowNr, column = 3).value = temp_group
          destinationSheet.cell(row = rowNr, column = 4).value = temp_title
          destinationSheet.cell(row = rowNr, column = 5).value = wms_layer
          destinationSheet.cell(row = rowNr, column = 6).value = wfs_layer
          destinationSheet.cell(row = rowNr, column = 7).value = temp_attribute
          if not temp_allowedRoles == "" and not temp_allowedRolesTemplate == "":
            destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles + ", " + temp_allowedRolesTemplate
          elif not temp_allowedRoles == "":
            destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRoles
          elif not temp_allowedRolesTemplate == "":
            destinationSheet.cell(row = rowNr, column = 8).value = temp_allowedRolesTemplate

          destinationSheet.cell(row = rowNr, column = 9).value = temp_serverType
          if temp_url == "":
            if defaultUrl == "":
              temp_url = "maps.metria.se/geoserver/"
            else:
              temp_url = defaultUrl 
          destinationSheet.cell(row = rowNr, column = 10).value = temp_url
          postNr = postNr + 1
          rowNr = rowNr + 1

          rowUpdated = 0
          nr_of_wms_layers = 0
          nr_of_wfs_layers = 0
          temp_group = ""
          temp_title = ""
          temp_serverType = ""
          temp_allowedRoles = ""
          temp_allowedRolesTemplate = ""
          temp_url = ""
          temp_attribute = ""
          splited_wfs_line = ""
          wms_layer = ""
          wfs_layer = ""

# ---------- Read information about Roles and update Excel ----------
    if '"allowedRolesTemplates":' in line:
      print("Reading roles ....")
      deleteEmptyRows(destinationFileName, wb, roleDescriptionSheet)
      roleDescriptionSheet.cell(row = 1, column = 1).font = Font(bold=True)
      roleDescriptionSheet.cell(row = 1, column = 1).value = "Nr"
      roleDescriptionSheet.cell(row = 1, column = 2).font = Font(bold=True)
      roleDescriptionSheet.cell(row = 1, column = 2).value = "Kund"
      roleDescriptionSheet.cell(row = 1, column = 3).font = Font(bold=True)
      roleDescriptionSheet.cell(row = 1, column = 3).value = "Rollgroup"
      roleDescriptionSheet.cell(row = 1, column = 4).font = Font(bold=True)
      roleDescriptionSheet.cell(row = 1, column = 4).value = "Roller"
      roleDetected = 1

      if roleDescriptionSheet.max_row > 1:
        roleRowNr = roleDescriptionSheet.max_row + 1
      else:
        roleRowNr = 2
    elif '},' in line and roleDetected == 1:
      roleDetected = 0

    elif roleDetected == 1:
      if '[' in line:
        role_group = line.replace('"', '')
        role_group = role_group.replace(':', '')
        role_group = role_group.replace('[', '')  
        role_group = role_group.strip()   
        roleRowUpdated = 1
      elif not ']' in line and roleRowUpdated == 1:
        temp_role = line.replace('"', '')
        temp_role = temp_role.strip()
        roleList.append(temp_role)
      elif ']' in line and roleRowUpdated == 1:
        roleDescriptionSheet.cell(row = roleRowNr, column = 1).alignment = Alignment(horizontal='left')
        roleDescriptionSheet.cell(row = roleRowNr, column = 1).value = rolePostNr
        roleDescriptionSheet.cell(row = roleRowNr, column = 2).value = customerName
        roleDescriptionSheet.cell(row = roleRowNr, column = 3).value = role_group
        tempRoleList = ""
        for x in roleList:
          if x == roleList[0]:
            tempRoleList = x
          else:
            tempRoleList = tempRoleList + " " + x
        roleDescriptionSheet.cell(row = roleRowNr, column = 4).value = tempRoleList
        roleRowNr = roleRowNr + 1
        rolePostNr = rolePostNr + 1
        roleRowUpdated = 0
        role_group = ""
        roleList = []

    line = f1.readline()

  f1.close()
  wb.save(destinationFileName)
  print(destinationFileName + " was updated!")

# -------- END ---------

